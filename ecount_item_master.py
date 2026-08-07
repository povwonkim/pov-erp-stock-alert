#!/usr/bin/env python3
"""이카운트 재고변동표(또는 품목등록) 엑셀에서 실제 품목마스터(품목코드↔브랜드↔조달유형)를
만들어 구글시트 `품목마스터` 탭에 반영한다.

핵심 아이디어(2026-07-28 확인): 이카운트 브랜드코드(품목그룹2코드)의 앞 2글자가 국가코드다
(KBP=KR013, HIGHTIDE=JP041, CAMBRIDGE IMPRINT=UK013 등). 이걸로 289개 브랜드를 사람이 일일이
안 보고도 자동 분류한다:
  - KR로 시작 → 국내사입(기본)
  - 그 외 국가코드(JP/US/UK/FR/DE...) → 해외수입(기본, 가장 보수적 = 안전한 기본값)
  - 숫자로만 된 코드(00001~00020대) → POV가 직접 관리하는 사내 카테고리. 이 중 실제로
    자체제작하는 브랜드만 SELF_MADE_BRANDS로 확정하고, 나머지는 AMBIGUOUS로 분류해
    조달유형은 임시로 해외수입 기본값을 쓰되 검토 리포트에 표로 뽑아준다.
  - MANUAL_OVERRIDES: 브랜드 원산지와 실제 조달경로가 다른 경우(PILOT/Midori는 일본
    브랜드지만 국내 유통사를 통해 사입) — 국가코드보다 우선 적용.

사용법:
  # 1) 분류 리포트만 보기(시트에 아무것도 안 씀)
  python3 ecount_item_master.py --source 재고변동표.xlsx --report-only

  # 2) 구글시트 품목마스터 탭에 실제 반영 (서버에서, gmail_token.json 필요)
  .venv/bin/python ecount_item_master.py --source 재고변동표.xlsx --spreadsheet-id <ID>
"""
from __future__ import annotations

import argparse
import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

# ---- 조달유형별 리드타임 (README/ecount_sheets_setup.py와 동일하게 유지) ----
# 국내위탁은 2026-08-07까지 이 표에도 classify()에도 없었다. 그런데 시트에는 국내위탁 40건이
# 들어있었다 — 코드가 만들 수 없는 값이니 사람이 시트에서 직접 고친 것이고, write_to_sheet가
# batchClear 후 전량 재작성이라 스크립트를 한 번만 다시 돌리면 그 40건이 전부 국내사입으로
# 덮여 사라진다. 조달유형을 정식으로 4종으로 늘려 그 경로를 막는다.
LEADTIME_BY_TYPE = {"자체제작": 30, "국내사입": 7, "국내위탁": 7, "해외수입": 21}

# 위탁 브랜드는 이카운트 품목그룹2명이 "위탁/ ALLGRAY"처럼 '위탁/' 접두어를 달고 있다
# (2026-08-07 실측: 위탁/ ALLGRAY 30건, 위탁/ JOPLINWORKS 5건, 위탁/ SOHOSU 3건,
# 위탁/ Cornervery 2건 = 40건, 전부 브랜드코드 KR014). 브랜드코드가 KR로 시작해서
# 국가코드 규칙으로는 국내사입으로 잘못 잡히므로 국가코드보다 먼저 판정해야 한다.
CONSIGNMENT_BRAND_PREFIX = "위탁/"

# 실제 POV 자체제작 브랜드로 확정된 것만 (2026-07-27~28 이카운트 품목그룹2 화면 + 사용자 확인).
# POV_application: 판매용은 아니지만(쇼핑백/봉투 등 비품) 재제작 필요 여부를 추적해야 해서
# 자체제작과 동일하게 관리.
SELF_MADE_BRANDS = {
    "POV_original", "POV Atelier", "POV x Hello Kitty", "POV x KBP", "POV_한글박물관",
    "POV_collaboration", "POV_inventario", "PRESSPRESS", "PRESSPRESS_collaboration", "양지사",
    "POV_application",
}

# 브랜드 원산지(국가코드)와 실제 조달경로가 다른 경우 — 국가코드 기본값보다 우선 적용.
# 2026-07-28 사용자 확인.
MANUAL_OVERRIDES = {
    "PILOT": "국내사입",             # 일본 브랜드, 국내 유통사 통해 사입
    "Midori": "국내사입",            # 동일
    "KBP": "국내사입",               # KR013 — 국가코드로도 이미 잡히지만 명시적으로 확인됨
    "GONGJANG(공장)": "국내사입",     # KR050 — 동일
    "BOOK": "국내사입",              # 매입
    "해외서적": "국내사입",           # 해외 서적이지만 국내에서 매입
    "VIEWPOINT": "국내사입",
    "프란츠(FRANZ)": "국내사입",
    "OTHER": "국내사입",             # 섞여있음 — SKU_OVERRIDES로 예외 처리(아래)
}

# 재고 성격이 안 맞거나(대여 서비스 등) 이 시스템에서 아예 다루지 않기로 한 브랜드.
# 2026-07-28 사용자 확인: 가져오지 않기로 함.
EXCLUDED_BRANDS = {"ARCHIVE. Object", "PointofView"}

# 브랜드 전체는 아니고 특정 품목만 제외 — 작년 시즌 아이템이라 재입고 안 함(2026-08-04
# 사용자 확인). 이카운트 품목명에 "(단종)"이 안 붙어있는 것들이라 브랜드/명칭 규칙으로는
# 못 걸러서 이름을 직접 등록한다. 품목명 앞뒤 공백은 비교 시 무시(ecount_daily_runner.py).
EXCLUDED_ITEM_NAMES = {
    "POV diary duo book 2026",
    "(50%)POV format series_calendar 2026",
    "(50%)POV format series_weekly 2026",
    "(50%)POV Original diary book 2026",
}

# 브랜드 단위로는 판단 안 되고 품목코드(SKU) 단위로 조달유형이 갈리는 예외.
# "OTHER" 브랜드 안에 섞여있는 것 중 확인된 것들(2026-07-28).
SKU_OVERRIDES = {
    "F-109": "해외수입",
    "IC005": "해외수입",
    "IT-VINTAGE-001": "국내사입",
    "891-10": "국내사입",
}

# 숫자 코드(POV 사내 카테고리)인데 위 규칙 어디에도 안 걸리는 것들 — 자동 분류 애매해서
# 사람 확인 전까지 임시로 해외수입(가장 보수적) 기본값을 쓴다. 확인되는 대로 여기 추가.
AMBIGUOUS_NUMERIC_BRANDS_DEFAULT = "해외수입"


def classify(brand: str, brand_code: str, item_code: str = "") -> tuple[str, bool]:
    """(조달유형, 확실함여부) 반환. 확실함=False면 리뷰 대상."""
    brand = (brand or "").strip()
    brand_code = (brand_code or "").strip()
    item_code = (item_code or "").strip()

    if item_code in SKU_OVERRIDES:
        return SKU_OVERRIDES[item_code], True
    if brand in SELF_MADE_BRANDS:
        return "자체제작", True
    if brand in MANUAL_OVERRIDES:
        return MANUAL_OVERRIDES[brand], True
    # 국가코드보다 먼저 — 위탁 브랜드는 브랜드코드가 KR014라 아래 규칙에 걸리면 국내사입이 된다.
    if brand.startswith(CONSIGNMENT_BRAND_PREFIX):
        return "국내위탁", True

    m = re.match(r"^([A-Za-z]{2})", brand_code)
    if m:
        country = m.group(1).upper()
        if country == "KR":
            return "국내사입", True
        return "해외수입", True  # 국가코드 있음 = 확정으로 취급(기본값이 곧 보수적 안전값)

    # 숫자 코드인데 자체제작 확정 리스트에 없음 — 애매함, 리뷰 필요.
    return AMBIGUOUS_NUMERIC_BRANDS_DEFAULT, False


def _clean(v) -> str:
    return (str(v).strip() if v is not None else "")


def load_rows(path: Path) -> list[dict]:
    """재고변동표(집계 모드) 엑셀에서 품목코드/품목명/브랜드/브랜드코드를 뽑는다.
    헤더 행 위치를 자동으로 찾는다(1행에 회사명/기간 안내가 있는 경우가 많음)."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_row = None
    for r in range(1, min(ws.max_row, 10) + 1):
        vals = [_clean(c.value) for c in ws[r]]
        if "품목코드" in vals and ("품목그룹2명" in vals or "브랜드" in vals):
            header_row = r
            break
    if header_row is None:
        raise SystemExit("헤더 행(품목코드/품목그룹2명)을 못 찾았습니다 — 파일 형식을 확인하세요.")

    headers = [_clean(c.value) for c in ws[header_row]]
    col = {h: i for i, h in enumerate(headers) if h}
    brand_col = col.get("품목그룹2명", col.get("브랜드"))
    brand_code_col = col.get("품목그룹2코드", col.get("브랜드코드"))
    code_col = col["품목코드"]
    name_col = col.get("품목명")

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        code = _clean(row[code_col]) if code_col < len(row) else ""
        if not code:
            continue
        rows.append({
            "품목코드": code,
            "품목명": _clean(row[name_col]) if name_col is not None and name_col < len(row) else "",
            "브랜드": _clean(row[brand_col]) if brand_col is not None and brand_col < len(row) else "",
            "브랜드코드": _clean(row[brand_code_col]) if brand_code_col is not None and brand_code_col < len(row) else "",
        })
    return rows


def build_item_master(rows: list[dict]) -> tuple[list[list], dict, int]:
    """품목마스터 행(품목마스터 탭 헤더 순서: 품목코드,품목명,브랜드,브랜드코드,조달유형,리드타임(일),갱신일)
    과 브랜드별 분류 리포트를 만든다. EXCLUDED_BRANDS는 아예 제외한다."""
    today = date.today().isoformat()
    out = []
    excluded_count = 0
    brand_stats: dict[str, dict] = {}

    for r in rows:
        if r["브랜드"] in EXCLUDED_BRANDS:
            excluded_count += 1
            continue
        ptype, certain = classify(r["브랜드"], r["브랜드코드"], r["품목코드"])
        out.append([
            r["품목코드"], r["품목명"], r["브랜드"], r["브랜드코드"],
            ptype, LEADTIME_BY_TYPE[ptype], today,
        ])
        s = brand_stats.setdefault(r["브랜드"], {"code": r["브랜드코드"], "types": {}, "certain": True, "count": 0})
        s["count"] += 1
        s["types"][ptype] = s["types"].get(ptype, 0) + 1
        if not certain:
            s["certain"] = False

    return out, brand_stats, excluded_count


def print_report(brand_stats: dict, excluded_count: int) -> None:
    by_type = {}
    for b, s in brand_stats.items():
        dominant = max(s["types"], key=s["types"].get)
        by_type.setdefault(dominant, []).append((b, s))

    print("\n[품목마스터] 조달유형별 브랜드/품목 집계:")
    for t in LEADTIME_BY_TYPE:
        items = by_type.get(t, [])
        total_skus = sum(s["count"] for _, s in items)
        print(f"  {t}: 브랜드 {len(items)}개, 품목 {total_skus}개")
    if excluded_count:
        print(f"  (제외됨: {excluded_count}개 품목 — {', '.join(EXCLUDED_BRANDS)})")

    mixed = [(b, s) for b, s in brand_stats.items() if len(s["types"]) > 1]
    if mixed:
        print(f"\n[품목마스터] 브랜드 내 SKU별로 조달유형이 갈리는 것 {len(mixed)}개:")
        for b, s in mixed:
            detail = ", ".join(f"{t}:{c}개" for t, c in s["types"].items())
            print(f"  - {b!r}: {detail}")

    ambiguous = [(b, s) for b, s in brand_stats.items() if not s["certain"]]
    if ambiguous:
        print(f"\n[품목마스터] ⚠️ 확인 필요(애매한 숫자코드 브랜드) {len(ambiguous)}개:")
        for b, s in sorted(ambiguous, key=lambda x: -x[1]["count"]):
            dominant = max(s["types"], key=s["types"].get)
            print(f"  - {b!r} (코드 {s['code']}, 품목 {s['count']}개) → 임시: {dominant}")


def merge_with_existing(rows: list[list], existing: list[list], *,
                        force_reclassify: bool = False) -> tuple[list[list], int]:
    """규칙이 새로 만든 rows에, 시트에 이미 있던 사람 손길(조달유형/리드타임)을 덮어씌운다.

    classify()는 브랜드코드 앞 2글자 국가코드로 조달유형을 '추정'할 뿐이라 틀릴 수 있고
    (PILOT/Midori가 일본 브랜드지만 국내사입인 게 대표 사례), 그래서 담당자가 시트에서
    직접 고친 값이 실제로 쌓여 있다. 2026-08-07 실측: 국내위탁 40건은 classify()가 반환조차
    할 수 없는 값이었고, 자체제작 626건의 리드타임 35일·국내사입 174건의 21일도 코드 기본값
    (30/7)과 다르다. 전부 사람이 고친 값이다.

    그래서 이미 시트에 있는 품목코드는 조달유형/리드타임을 건드리지 않는다. 규칙 개선을
    소급 적용하고 싶을 때만 --force-reclassify로 덮어쓴다.
    반환: (병합된 행, 보존된 건수)
    """
    # 기존 행: [품목코드, 품목명, 브랜드, 브랜드코드, 조달유형, 리드타임(일), 갱신일]
    kept: dict[str, tuple[str, str]] = {}
    for r in existing:
        if not r or not str(r[0]).strip():
            continue
        ptype = str(r[4]).strip() if len(r) > 4 else ""
        # '미분류'는 사람이 아직 확인 안 한 자리표시자라 보존 대상이 아니다 — 규칙이 다시 채운다.
        if ptype and ptype != "미분류":
            kept[str(r[0]).strip()] = (ptype, str(r[5]).strip() if len(r) > 5 else "")

    if force_reclassify:
        return rows, 0

    def same_leadtime(a, b) -> bool:
        # 시트에서 읽으면 '21.0', 코드가 만들면 21 — 표기만 다르고 같은 값인 경우가 많다.
        try:
            return float(a) == float(b)
        except (TypeError, ValueError):
            return str(a).strip() == str(b).strip()

    preserved = 0
    merged = []
    for row in rows:
        code = str(row[0]).strip()
        if code in kept:
            ptype, leadtime = kept[code]
            if row[4] != ptype or (leadtime and not same_leadtime(row[5], leadtime)):
                preserved += 1
            row = list(row)
            row[4] = ptype
            if leadtime:
                row[5] = leadtime
        merged.append(row)
    return merged, preserved


def write_to_sheet(spreadsheet_id: str, rows: list[list], *, force_reclassify: bool = False) -> None:
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    from ecount_sheets_setup import _TOKEN_FILE, SCOPES, DATA_START_IDX, ITEM_MASTER_TAB

    creds = Credentials.from_authorized_user_file(str(_TOKEN_FILE), SCOPES)
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    service = build("sheets", "v4", credentials=creds)

    # 덮어쓰기 전에 지금 시트에 뭐가 있는지 먼저 읽는다 — 담당자가 손으로 고친 조달유형/
    # 리드타임을 그대로 살려서 다시 써야 하기 때문. 이걸 안 하면 이 스크립트를 한 번 돌리는
    # 것만으로 사람이 쌓아둔 판단이 전부 규칙 기본값으로 리셋된다.
    existing = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=f"'{ITEM_MASTER_TAB}'!A{DATA_START_IDX + 1}:G100000",
    ).execute().get("values", [])
    rows, preserved = merge_with_existing(rows, existing, force_reclassify=force_reclassify)
    if force_reclassify:
        print(f"[품목마스터] ⚠️ --force-reclassify — 기존 {len(existing)}행의 사람 수정값을 규칙값으로 덮어씁니다")
    else:
        print(f"[품목마스터] 기존 {len(existing)}행 중 사람이 고친 조달유형/리드타임 {preserved}건 보존")

    service.spreadsheets().values().batchClear(
        spreadsheetId=spreadsheet_id, body={"ranges": [f"'{ITEM_MASTER_TAB}'!A{DATA_START_IDX + 1}:Z100000"]}
    ).execute()
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"'{ITEM_MASTER_TAB}'!A{DATA_START_IDX + 1}",
        valueInputOption="RAW",
        body={"values": rows},
    ).execute()
    print(f"[품목마스터] 구글시트 반영 완료: {len(rows)}행")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", required=True, help="재고변동표(집계) 엑셀 경로")
    ap.add_argument("--spreadsheet-id", help="지정하면 구글시트 품목마스터 탭에 반영")
    ap.add_argument("--report-only", action="store_true", help="시트에 안 쓰고 분류 리포트만 출력")
    ap.add_argument("--force-reclassify", action="store_true",
                    help="사람이 시트에서 고친 조달유형/리드타임까지 규칙값으로 덮어쓴다 (기본은 보존)")
    args = ap.parse_args()

    rows = load_rows(Path(args.source))
    print(f"[품목마스터] {len(rows)}개 품목 로드")

    master_rows, brand_stats, excluded_count = build_item_master(rows)
    print_report(brand_stats, excluded_count)

    if args.report_only or not args.spreadsheet_id:
        print("\n[품목마스터] --report-only 또는 --spreadsheet-id 미지정 — 시트에 쓰지 않음")
        return 0

    write_to_sheet(args.spreadsheet_id, master_rows, force_reclassify=args.force_reclassify)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
