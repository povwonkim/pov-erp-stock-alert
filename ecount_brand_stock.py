#!/usr/bin/env python3
"""POV 자체제작 브랜드(품목그룹2) 재고 트래킹.

이카운트 "재고변동표"(또는 "품목등록") 화면에서 내보낸 엑셀을 읽어
POV 자체제작 브랜드만 필터링하고, 저재고 항목을 표시한다.
API로는 입고수량/출고수량을 못 가져오므로(이카운트 OAPI에 재고변동
조회 엔드포인트 자체가 없음 — 확인 완료), 사람이 이카운트에서 내려받은
엑셀을 입력으로 받는다.

입력 엑셀 기대 컬럼 (헤더로 매칭, 순서 무관):
  품목그룹2명 / 품목그룹2코드 / 품목코드 / 품목명 /
  전일재고 / 입고수량 / 출고수량 / 재고수량 / 입고단가

사용법:
  python3 ecount_brand_stock.py --source 재고변동표.xlsx
  python3 ecount_brand_stock.py --source 재고변동표.xlsx --threshold 10 --out cron_tracking/brand/latest.json
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# POV 자체제작 브랜드 (품목그룹2명 기준). 이카운트 품목그룹2 화면에서 확인한 값.
POV_BRANDS = {
    "POV_original",
    "POV Atelier",
    "POV x Hello Kitty",
    "POV x KBP",
    "POV_한글박물관",
    "POV_collaboration",
    "POV_inventario",
    "PRESSPRESS",
    "PRESSPRESS_collaboration",
    "양지사",
}

# 엑셀 헤더 -> 내부 필드명. 이카운트 화면 버전에 따라 표기가 조금씩 다를 수 있어 후보를 둔다.
HEADER_ALIASES = {
    "브랜드": ["품목그룹2명", "품목그룹2"],
    "브랜드코드": ["품목그룹2코드"],
    "품목코드": ["품목코드", "PROD_CD"],
    "품목명": ["품목명", "PROD_DES"],
    "전일재고": ["전일재고"],
    "입고수량": ["입고수량"],
    "출고수량": ["출고수량"],
    "재고수량": ["재고수량"],
    "입고단가": ["입고단가"],
}


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    """헤더 행과 {내부필드명: 열번호} 매핑을 찾는다."""
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        cells = [c.value for c in ws[row_idx]]
        col_map: dict[str, int] = {}
        for field, aliases in HEADER_ALIASES.items():
            for col_idx, val in enumerate(cells):
                if isinstance(val, str) and val.strip() in aliases:
                    col_map[field] = col_idx
                    break
        # 최소한 품목코드 + 브랜드는 찾아야 헤더 행으로 인정
        if "품목코드" in col_map and "브랜드" in col_map:
            return row_idx, col_map
    raise SystemExit("헤더 행을 찾지 못했습니다 — 엑셀에 '품목코드'와 '품목그룹2명' 컬럼이 있는지 확인하세요.")


def _to_number(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return 0.0


def load_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_row, col_map = _find_header_row(ws)

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        prod_cd = row[col_map["품목코드"]] if col_map.get("품목코드") is not None else None
        if not prod_cd:
            continue
        rec = {
            "브랜드": row[col_map["브랜드"]] if "브랜드" in col_map else "",
            "브랜드코드": row[col_map["브랜드코드"]] if "브랜드코드" in col_map else "",
            "품목코드": str(prod_cd).strip(),
            "품목명": row[col_map["품목명"]] if "품목명" in col_map else "",
            "전일재고": _to_number(row[col_map["전일재고"]]) if "전일재고" in col_map else None,
            "입고수량": _to_number(row[col_map["입고수량"]]) if "입고수량" in col_map else None,
            "출고수량": _to_number(row[col_map["출고수량"]]) if "출고수량" in col_map else None,
            "재고수량": _to_number(row[col_map["재고수량"]]) if "재고수량" in col_map else None,
            "입고단가": _to_number(row[col_map["입고단가"]]) if "입고단가" in col_map else None,
        }
        rows.append(rec)
    return rows


def filter_pov_brands(rows: list[dict]) -> list[dict]:
    return [r for r in rows if str(r.get("브랜드", "")).strip() in POV_BRANDS]


def flag_low_stock(rows: list[dict], threshold: float) -> list[dict]:
    out = []
    for r in rows:
        qty = r.get("재고수량")
        r = dict(r)
        r["저재고"] = qty is not None and qty <= threshold
        out.append(r)
    return out


def summarize_by_brand(rows: list[dict]) -> dict[str, dict]:
    summary: dict[str, dict] = {}
    for r in rows:
        b = r.get("브랜드") or "(미분류)"
        s = summary.setdefault(b, {"품목수": 0, "총재고수량": 0.0, "저재고품목수": 0})
        s["품목수"] += 1
        s["총재고수량"] += r.get("재고수량") or 0.0
        if r.get("저재고"):
            s["저재고품목수"] += 1
    return summary


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", required=True, help="이카운트에서 내보낸 재고변동표/품목등록 엑셀 경로")
    ap.add_argument("--threshold", type=float, default=10, help="저재고로 표시할 재고수량 임계값 (기본 10)")
    ap.add_argument("--out", help="결과 JSON 저장 경로 (미지정 시 콘솔 요약만 출력)")
    ap.add_argument("--all-brands", action="store_true", help="POV 자체제작 브랜드 필터를 걸지 않고 전체 출력")
    args = ap.parse_args()

    rows = load_rows(Path(args.source))
    print(f"[brand] 엑셀에서 {len(rows)}개 품목 로드")

    target = rows if args.all_brands else filter_pov_brands(rows)
    if not args.all_brands:
        print(f"[brand] POV 자체제작 브랜드 필터 적용 → {len(target)}개 품목")

    target = flag_low_stock(target, args.threshold)
    summary = summarize_by_brand(target)

    print("\n[brand] 브랜드별 요약:")
    for brand, s in sorted(summary.items()):
        print(f"  - {brand}: 품목 {s['품목수']}개, 총재고 {s['총재고수량']:.0f}, 저재고 {s['저재고품목수']}개")

    low = [r for r in target if r.get("저재고")]
    if low:
        print(f"\n[brand] 저재고(≤{args.threshold}) 품목 {len(low)}건:")
        for r in sorted(low, key=lambda r: (r["브랜드"], r.get("재고수량") or 0)):
            print(f"  - [{r['브랜드']}] {r['품목코드']} {r['품목명']} → 재고 {r.get('재고수량')}")

    if args.out:
        out_path = Path(args.out)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(
            json.dumps({"summary": summary, "items": target}, ensure_ascii=False, indent=2)
        )
        print(f"\n[brand] 결과 저장: {out_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
