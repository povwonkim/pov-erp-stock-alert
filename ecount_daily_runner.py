#!/usr/bin/env python3
"""오프라인 재고 감시 시스템 — 매일 자동 실행되는 통합 스크립트.

흐름:
  1. 이카운트 창고별재고현황 API로 오늘 시점 재고 스냅샷 조회 (전체 창고, RAW 그대로 저장)
  2. Gmail에서 이카운트 "판매현황" 자동발송 이메일의 첨부 엑셀을 받아 파싱 (전체 창고, RAW 그대로 저장)
  3. 두 원본을 오프라인 4개 창고로 필터링 + 품목마스터(브랜드/조달유형/리드타임) 조인
  4. 일별재고이력에서 전일 데이터를 읽어 전일재고·품절경과일 연속성 계산, 입고수량 역산
  5. DOI·상태·우선순위·조치방안 계산 (README "DOI 기반 우선순위 체계" 참고)
  6. 일별재고이력에 오늘자 행 누적, 3층 결과물 탭(디자인팀_발주필요/관리팀_전체재고/악성재고/
     악성품절) 재작성

기준일(TARGET_DATE) 하나로 통일: 이카운트 판매현황 자동알림의 "기준일자=전일" 관례를 그대로
따라, 이 스크립트가 실행되는 날의 "어제"를 재고/판매 양쪽의 기준일로 쓴다.

사용법 (서버에서):
  .venv/bin/python ecount_daily_runner.py --spreadsheet-id <ID>
  .venv/bin/python ecount_daily_runner.py --spreadsheet-id <ID> --sales-xlsx cron_tracking/ecount/sales.xlsx
  .venv/bin/python ecount_daily_runner.py --spreadsheet-id <ID> --dry-run   # 시트에 안 쓰고 요약만 출력
"""
from __future__ import annotations

import argparse
import json
import time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

from ecount_client import EcountClient
from ecount_item_master import EXCLUDED_BRANDS, EXCLUDED_ITEM_NAMES, LEADTIME_BY_TYPE
from ecount_sheets_setup import (
    TABS, OFFLINE_WAREHOUSES, OFFLINE_WAREHOUSE_CODES, DATA_START_IDX, _TOKEN_FILE, SCOPES,
    STATUS_COLOR_RULES, BANNER1_BG, BANNER_FG_LIGHT, BANNER3_BG, BANNER3_FG,
    HEADER_BG, HEADER_FG, FONT_BODY, FONT_MONO, BODY_FG,
    SEMANTIC_DANGER_BG, SEMANTIC_INFO_BG, SEMANTIC_WARNING_BG,
    NUMBER_FORMAT_MONEY,
)

KST = timezone(timedelta(hours=9))
DUMP_DIR = Path(__file__).parent / "cron_tracking" / "ecount"

# 품목코드 → 입고단가 조회표. 이카운트 재고 API(fetch_inventory_raw)엔 단가 정보가 없어서,
# 2026-07-30에 사용자가 뽑아준 재고현황 엑셀에서 1회성으로 추출했다(8,399개 품목,
# one_time_seed_2026-07-30/item_cost_lookup.json). 입고단가는 SKU당 잘 안 바뀌는 값이라
# 당분간은 이 스냅샷으로 충분하지만, 이카운트 API에서 직접 가져오는 방법이 생기면 교체할 것
# (2026-08-06 사용자 요청 — 악성재고/샘플의심재고에 재고금액 표시).
_COST_LOOKUP_PATH = Path(__file__).parent / "one_time_seed_2026-07-30" / "item_cost_lookup.json"
try:
    ITEM_COST_LOOKUP: dict[str, float] = json.loads(_COST_LOOKUP_PATH.read_text())
except FileNotFoundError:
    ITEM_COST_LOOKUP = {}

# 이카운트 실서버 조회 API(재고현황/창고별재고현황 포함)는 종류당 1회/10분 제한(2026-07-28
# 공식 문서 확인, HTTP 412 = "API 전송 횟수 기준을 넘은 경우"). 오프라인 창고 4곳을 각각
# 조회하려면 그만큼 간격을 둬야 한다 — 10분(600초) + 여유.
INVENTORY_CALL_INTERVAL_SEC = 610

# DOI 위험/주의 임계값 (일) — README "DOI 기반 우선순위 체계" 표와 동일.
# 자체제작 리드타임 35일 → 30일로 변경(2026-07-28 사용자 확정, ecount_item_master.py의
# LEADTIME_BY_TYPE도 동일하게 맞춤).
RISK_WARN_BY_TYPE = {
    "자체제작": (30, 44),
    "국내사입": (7, 14),
    "해외수입": (21, 35),
}

# 판매로 안 잡히는 비품 브랜드는 DOI(판매속도 기반) 로직이 원천적으로 안 맞는다 — 고객에게
# 파는 게 아니라 소모되기만 해서 최근7일 판매량이 항상 0으로 잡히고, 그러면 determine_status()가
# 위험/주의/과잉을 절대 못 만난다(재고>5면 그냥 계속 정상). 대신 단순 재고 임계값으로 판정한다.
# 2026-07-28 사용자 확정: POV_application(쇼핑백/봉투 등) 제작도 ~1개월 걸려 재고 1000개 이하면
# 재발주 시점(품절/마이너스재고는 그대로 determine_status()가 판정).
NON_SALES_REORDER_THRESHOLD = {"POV_application": 1000}

# 2026-07-28 사용자 확정: 품절-신규(0~9일)/품절-지속(10~29일) 2단계가 표에서 흩어져 보여
# 헷갈린다는 피드백 → 품절-지속 하나로 합치고(0~29일), 마이너스재고는 위험과 헷갈리지 않게
# 이모지를 ⛔️로 분리.
PRIORITY_BY_STATUS = {
    "🔴 위험": 1, "🟤 품절-지속": 1,
    "🟠 주의": 2,
    "🟡 재고소량": 3,
    "⚫ 품절-장기": 3,
}

ACTION_BY_STATUS = {
    "⛔️ 마이너스재고": "재고 데이터 확인 필요",
    "🔴 위험": "긴급 제작 필요",
    "🟤 품절-지속": "긴급 제작 필요",
    "🟠 주의": "제작 검토",
    "🟡 재고소량": "마케팅 부스트 또는 정리 검토",
    "🔵 과잉": "프로모션/할인 검토",
    "🟢 정상": "-",
}


# ---------------------------------------------------------------------------
# 1. 이카운트 재고 API
# ---------------------------------------------------------------------------

def fetch_inventory_raw(base_date: str) -> list[dict]:
    """창고별재고현황 API 원본 — 오프라인 4개 창고만, 창고별로 나눠서 호출.

    전체 창고를 한 번에 조회하면 10000건 근처에서 잘리는 것으로 의심되는 문제가 있었다
    (README '탐사 결과' 참고). WH_CD로 창고를 하나씩 지정해 호출하면 건수가 훨씬 적게
    나오는 것을 실제로 확인(2026-07-28, WH_CD=00014 단독 6676건 < 전체 10000건) —
    이 시스템은 애초에 오프라인 4개 창고 외엔 안 쓰므로, 온라인 창고는 아예 조회하지 않는다.

    단, 실서버 조회 API는 종류당 1회/10분 제한이 있어(공식 문서 확인, 2026-07-28) 창고 4개를
    연달아 부르면 두 번째부터 무조건 막힌다. 그래서 호출 사이에 실제로 대기한다 — 총 소요시간이
    30~40분이지만, 하루 한 번 도는 배치라 문제없다.
    """
    client = EcountClient()
    rows = []
    for i, (wh_name, wh_code) in enumerate(OFFLINE_WAREHOUSE_CODES.items()):
        if i > 0:
            print(f"[runner] 조회 API 1회/10분 제한 — 다음 창고까지 {INVENTORY_CALL_INTERVAL_SEC}초 대기...")
            time.sleep(INVENTORY_CALL_INTERVAL_SEC)
        print(f"[runner] {wh_name}(WH_CD={wh_code}) 조회 중...")
        data = client.inventory_balance_by_location(base_date, WH_CD=wh_code)
        result = data.get("Data", {}).get("Result") or []
        if len(result) >= 10000:
            print(f"[runner] ⚠️ {wh_name}(WH_CD={wh_code}) 응답이 {len(result)}건 — 10000건 "
                  "근처라 이 창고도 페이지네이션으로 잘렸을 가능성이 있습니다. 확인 필요.")
        for r in result:
            try:
                qty = float(r.get("BAL_QTY") or 0)
            except ValueError:
                qty = 0.0
            rows.append({
                "창고코드": r.get("WH_CD", wh_code),
                "창고명": r.get("WH_DES", wh_name),
                "품목코드": r.get("PROD_CD", ""),
                "품목명": r.get("PROD_DES", ""),
                "사이즈": r.get("PROD_SIZE_DES", ""),
                "재고수량": qty,
            })
    return rows


# ---------------------------------------------------------------------------
# 2. Gmail 판매현황 첨부 엑셀
# ---------------------------------------------------------------------------

def fetch_sales_xlsx_from_gmail(query: str, out_path: Path) -> Path:
    from ecount_gmail_fetch import _load_creds, list_messages, download_first_attachment
    from googleapiclient.discovery import build

    creds = _load_creds()
    service = build("gmail", "v1", credentials=creds)
    messages = list_messages(service, query)
    if not messages:
        raise SystemExit(f"[runner] Gmail 쿼리에 매치되는 메일이 없습니다: {query!r}")
    saved = download_first_attachment(service, messages[0]["id"], out_path)
    if not saved:
        raise SystemExit("[runner] 최신 메일에서 엑셀 첨부를 못 찾았습니다.")
    return saved


def _clean(v) -> str:
    return (str(v).strip() if v is not None else "")


def _to_number(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return 0.0


SALES_HEADER_MAP = {
    "품목그룹2명": "브랜드", "브랜드": "브랜드",
    "품목코드": "품목코드",
    "품명 및 규격": "품명", "품명": "품명",
    "수량": "수량", "단가": "단가", "공급가액": "공급가액", "부가세": "부가세",
    "합계": "합계", "적요": "적요", "창고명": "창고명",
    "사원(담당)명": "담당자", "담당자": "담당자",
}


def parse_sales_xlsx(path: Path) -> list[dict]:
    """판매현황 엑셀(RAW_판매현황 원본) 파싱. 헤더 행을 자동으로 찾는다."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_row = None
    for r in range(1, min(ws.max_row, 10) + 1):
        vals = [_clean(c.value) for c in ws[r]]
        if "품목코드" in vals and "수량" in vals:
            header_row = r
            break
    if header_row is None:
        raise SystemExit("[runner] 판매현황 엑셀에서 헤더 행(품목코드/수량)을 못 찾았습니다.")

    headers = [_clean(c.value) for c in ws[header_row]]
    col = {SALES_HEADER_MAP[h]: i for i, h in enumerate(headers) if h in SALES_HEADER_MAP}
    if "품목코드" not in col:
        raise SystemExit("[runner] 판매현황 엑셀에 품목코드 컬럼이 없습니다.")

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        def get(field, default=""):
            idx = col.get(field)
            return row[idx] if idx is not None and idx < len(row) else default

        code = _clean(get("품목코드"))
        if not code:
            continue
        rows.append({
            "브랜드": _clean(get("브랜드")),
            "품목코드": code,
            "품명": _clean(get("품명")),
            "수량": _to_number(get("수량")),
            "단가": _to_number(get("단가")),
            "공급가액": _to_number(get("공급가액")),
            "부가세": _to_number(get("부가세")),
            "합계": _to_number(get("합계")),
            "적요": _clean(get("적요")),
            "창고명": _clean(get("창고명")),
            "담당자": _clean(get("담당자")),
        })
    return rows


# ---------------------------------------------------------------------------
# 3. 구글시트 읽기/쓰기 helpers
# ---------------------------------------------------------------------------

def _load_creds():
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials

    creds = Credentials.from_authorized_user_file(str(_TOKEN_FILE), SCOPES)
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
        _TOKEN_FILE.write_text(creds.to_json())
    return creds


def read_tab_rows(service, spreadsheet_id: str, tab_name: str) -> list[dict]:
    headers = TABS[tab_name]["headers"]
    last_col = chr(ord("A") + len(headers) - 1) if len(headers) <= 26 else "Z"
    rng = f"'{tab_name}'!A{DATA_START_IDX + 1}:{last_col}200000"
    # UNFORMATTED_VALUE 필수 — 안 그러면 숫자서식이 적용된 표시문자열("35일", "–")이 그대로
    # 와서 파싱이 깨진다 (예: 품절경과일=0인 셀은 서식상 "–"로 보이지만 실제 값은 0이어야 함).
    resp = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id, range=rng, valueRenderOption="UNFORMATTED_VALUE"
    ).execute()
    values = resp.get("values", [])
    out = []
    for row in values:
        row = row + [""] * (len(headers) - len(row))
        out.append(dict(zip(headers, row)))
    return out


def replace_tab_rows(service, spreadsheet_id: str, tab_name: str, rows: list[list]) -> None:
    service.spreadsheets().values().batchClear(
        spreadsheetId=spreadsheet_id, body={"ranges": [f"'{tab_name}'!A{DATA_START_IDX + 1}:Z200000"]}
    ).execute()
    if not rows:
        return
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"'{tab_name}'!A{DATA_START_IDX + 1}",
        valueInputOption="RAW",
        body={"values": rows},
    ).execute()


def append_history_rows(service, spreadsheet_id: str, target_date_str: str, new_rows: list[list]) -> None:
    """일별재고이력은 누적이지만, 같은 날짜로 재실행하면 그 날짜 행만 지우고 다시 쓴다(재실행 안전)."""
    headers = TABS["일별재고이력"]["headers"]
    existing = read_tab_rows(service, spreadsheet_id, "일별재고이력")
    kept = [[r[h] for h in headers] for r in existing if r["날짜"] != target_date_str]
    all_rows = kept + new_rows
    replace_tab_rows(service, spreadsheet_id, "일별재고이력", all_rows)


def auto_register_new_items(item_master: dict, sales_raw: list[dict], target_date: date) -> list[list]:
    """오늘 판매현황에 새로 나타난(품목마스터에 없는) 품목코드를 자동 등록한다.

    조달유형은 SKU가 아니라 브랜드 단위 속성이므로(README 참고), 같은 브랜드의 기존
    품목마스터 항목에서 다수결로 물려받는다 — 사람이 매번 새 상품을 등록할 필요가 없다.
    브랜드 자체가 처음 등장하는 경우만 '미분류'로 남아 사람이 한 번 확인하면 되고, 그 뒤로는
    같은 브랜드의 모든 신상품에 자동 적용된다(이카운트 원본에 브랜드코드가 없어 국가코드
    휴리스틱은 못 쓰지만, 브랜드명 다수결만으로 충분).

    item_master는 in-place로 갱신되어 오늘 계산부터 바로 반영되고, 반환값은 품목마스터
    시트에 추가로 써야 할 행이다(기존 행은 건드리지 않고 append).
    """
    brand_type_votes: dict[str, dict[str, int]] = {}
    for meta in item_master.values():
        if not meta.get("브랜드") or not meta.get("조달유형") or meta["조달유형"] == "미분류":
            continue
        votes = brand_type_votes.setdefault(meta["브랜드"], {})
        votes[meta["조달유형"]] = votes.get(meta["조달유형"], 0) + 1

    new_rows: list[list] = []
    seen_codes: set[str] = set()
    for r in sales_raw:
        code = r["품목코드"]
        if not code or code in item_master or code in seen_codes:
            continue
        seen_codes.add(code)
        brand = r["브랜드"]
        votes = brand_type_votes.get(brand)
        if votes:
            ptype = max(votes, key=votes.get)
            leadtime = LEADTIME_BY_TYPE.get(ptype, "")
        else:
            ptype, leadtime = "미분류", ""
        item_master[code] = {"브랜드": brand, "조달유형": ptype, "리드타임": leadtime, "품목명": r["품명"]}
        new_rows.append([code, r["품명"], brand, "", ptype, leadtime, target_date.isoformat()])
    return new_rows


def append_item_master_rows(service, spreadsheet_id: str, rows: list[list]) -> None:
    if not rows:
        return
    service.spreadsheets().values().append(
        spreadsheetId=spreadsheet_id,
        range=f"'품목마스터'!A{DATA_START_IDX + 1}",
        valueInputOption="RAW",
        insertDataOption="INSERT_ROWS",
        body={"values": rows},
    ).execute()


# ---------------------------------------------------------------------------
# 4. 계산 로직
# ---------------------------------------------------------------------------

def determine_status(재고: float, 최근7일: float, 품절경과일: int, 조달유형: str | None) -> str:
    # 상태 이모지는 상태별 개별 아이콘이 아니라 동그라미로 통일한다(2026-07-28, 전사 배포용
    # "읽는 법" 탭 도입과 함께 확정, 같은 날 재조정): ⛔️ 마이너스재고(위험과 헷갈리지 않게 분리) ·
    # 🔴 위험 · 🟠 주의 · 🟡 재고소량 · 🔵 과잉 · 🟢 정상 · 🟤 품절-지속(0~29일, 원래 품절-신규/지속
    # 2단계였는데 표에서 흩어져 헷갈린다는 피드백으로 통합) · ⚫ 품절-장기(30일+).
    # STATUS_COLOR_RULES의 조건부서식은 이모지가 아니라 한글 부분일치라 영향 없음.
    if 재고 < 0:
        return "⛔️ 마이너스재고"
    if 재고 == 0:
        if 품절경과일 <= 29:
            return "🟤 품절-지속"
        return "⚫ 품절-장기"
    if 재고 <= 5 and 최근7일 == 0:
        return "🟡 재고소량"
    if 최근7일 > 0:
        doi = 재고 / (최근7일 / 7)
        risk, warn = RISK_WARN_BY_TYPE.get(조달유형 or "", (None, None))
        if risk is not None:
            if doi <= risk:
                return "🔴 위험"
            if doi <= warn:
                return "🟠 주의"
        if doi > 180:
            return "🔵 과잉"
        return "🟢 정상"
    return "🟢 정상"


def compute_doi(재고: float, 최근7일: float) -> float | str:
    # 재고 0 이하(품절/마이너스재고)일 때도 빈칸 — 안 그러면 마이너스재고(음수 재고)에서 DOI가 음수로
    # 나와 의미 없는 값이 찍힌다 (설계 스펙 "POV_재고관리_마스터_설계.md" 명시 사항).
    if 최근7일 <= 0 or 재고 <= 0:
        return ""
    return round(재고 / (최근7일 / 7), 1)


# 악성품절 세분화 구간(2026-08-04 사용자 확정) — 90일 미만은 악성품절 대상에서 아예 제외.
STOCKOUT_BUCKETS = [
    (730, "2년+"), (365, "1년+"), (180, "180일+"), (150, "150일+"), (120, "120일+"), (90, "90일+"),
]


def stockout_bucket(품절경과일: int) -> str:
    for threshold, label in STOCKOUT_BUCKETS:
        if 품절경과일 >= threshold:
            return label
    return ""


def build_daily_rows(target_date: date, inventory_raw: list[dict], sales_raw: list[dict],
                      item_master: dict[str, dict], history: list[dict]) -> dict:
    """오늘자 일별재고이력 행들과 3층 결과물을 계산해서 반환."""
    target_str = target_date.isoformat()
    prev_str = (target_date - timedelta(days=1)).isoformat()

    # 오프라인 4개 창고로 필터링한 재고 — 품목코드별 총재고 + 창고별 재고.
    offline_inv = [r for r in inventory_raw if r["창고명"] in OFFLINE_WAREHOUSES]
    total_by_item: dict[str, float] = {}
    by_item_wh: dict[str, dict[str, float]] = {}
    name_by_item: dict[str, str] = {}
    for r in offline_inv:
        code = r["품목코드"]
        total_by_item[code] = total_by_item.get(code, 0.0) + r["재고수량"]
        by_item_wh.setdefault(code, {})
        by_item_wh[code][r["창고명"]] = by_item_wh[code].get(r["창고명"], 0.0) + r["재고수량"]
        if r["품목명"]:
            name_by_item[code] = r["품목명"]

    # 오프라인 4개 창고로 필터링한 판매(출고) — 품목코드별 합계.
    offline_sales = [r for r in sales_raw if r["창고명"] in OFFLINE_WAREHOUSES]
    out_by_item: dict[str, float] = {}
    for r in offline_sales:
        code = r["품목코드"]
        out_by_item[code] = out_by_item.get(code, 0.0) + r["수량"]
        if r["품명"] and code not in name_by_item:
            name_by_item[code] = r["품명"]

    # 이력 인덱싱 — 품목코드별 (날짜 -> row).
    hist_by_item: dict[str, dict[str, dict]] = {}
    for r in history:
        hist_by_item.setdefault(r["품목코드"], {})[r["날짜"]] = r

    # 처리 대상 품목 = 품목마스터 전체(관리팀 전체 그림) ∪ 오늘 재고/판매에 등장한 품목(미분류 포함).
    all_codes = set(item_master) | set(total_by_item) | set(out_by_item)

    history_rows: list[list] = []
    design_rows: list[list] = []
    mgmt_rows: list[list] = []
    malstock_rows: list[list] = []
    maldead_rows: list[list] = []
    sample_rows: list[list] = []

    for code in sorted(all_codes):
        meta = item_master.get(code, {})
        브랜드 = meta.get("브랜드", "미분류")
        조달유형 = meta.get("조달유형", "미분류")
        리드타임 = meta.get("리드타임", "")
        품목명 = name_by_item.get(code, meta.get("품목명", ""))

        # 단종/제외 품목 — 이 시스템 어디에도(일별재고이력 포함) 아예 안 남긴다(2026-08-04
        # 사용자 확정: "시즌아이템이라서 이제는 단종되거나 여기에 두어도 의미없는 건 제외").
        # 품목명에 "(단종)"이 붙어있으면 이카운트 쪽에서 이미 단종 처리한 것으로 보고 제외.
        # EXCLUDED_BRANDS(ARCHIVE. Object/PointofView)는 2026-07-28에 이미 제외하기로
        # 확정했었는데 이 파이프라인엔 안 걸려있었던 걸 여기서 같이 바로잡는다.
        # 특정 품목 개별 제외(작년 시즌 아이템 등, 재입고 안 함 — 2026-08-04 사용자 확인).
        # 품목명 앞뒤 공백 차이로 안 걸러지는 일이 없게 strip 비교.
        if "(단종)" in 품목명 or 브랜드 in EXCLUDED_BRANDS or 품목명.strip() in EXCLUDED_ITEM_NAMES:
            continue

        재고 = total_by_item.get(code, 0.0)
        출고 = out_by_item.get(code, 0.0)
        item_hist_so_far = hist_by_item.get(code, {})
        prev_row = item_hist_so_far.get(prev_str)
        # 전일 기록이 없는데(prev_row=None) 이 품목의 더 이전 이력은 있다면 — 진짜 0에서
        # 시작하는 게 아니라 그 사이 실행이 끊겨서 생긴 구멍이다(2026-08-04 확인: 크론이
        # 며칠 실패하던 시기에 이런 구멍이 실제로 생겼고, 전일재고를 0으로 잘못 가정해서
        # 그 차이가 전부 "입고"로 잡히는 아티팩트가 났었다 — fix_cold_start_artifacts.py로
        # 기존 데이터는 정리했고, 여기서는 재발을 막는다). 이럴 땐 입고를 계산하지 않고
        # 빈칸으로 남겨 "모름"을 정직하게 표시한다.
        is_gap = prev_row is None and bool(item_hist_so_far)
        전일재고 = _to_number(prev_row["재고"]) if prev_row else 0.0
        입고계산 = "" if is_gap else (재고 - 전일재고) + 출고

        prev_stockout_days = int(_to_number(prev_row["품절(일)"])) if prev_row and prev_row.get("품절(일)") not in ("", None) else 0
        if 재고 <= 0:
            # prev_row가 아예 없는(처음 등장하는) 품목은 "어제도 품절이었다"고 이어받을 근거가
            # 없으므로 무조건 0일차(처음 확인)로 시작한다.
            품절경과일 = (prev_stockout_days + 1) if (prev_row is not None and 전일재고 <= 0) else 0
        else:
            품절경과일 = 0

        # 최근 7일/90일 판매량 = 과거 이력(오늘 제외) + 오늘 출고.
        item_hist = hist_by_item.get(code, {})
        최근7일 = 출고
        최근90일 = 출고
        for delta in range(1, 90):
            d = (target_date - timedelta(days=delta)).isoformat()
            row = item_hist.get(d)
            if not row:
                continue
            qty = _to_number(row.get("출고"))
            if delta <= 6:
                최근7일 += qty
            최근90일 += qty

        reorder_threshold = NON_SALES_REORDER_THRESHOLD.get(브랜드)
        if reorder_threshold is not None and 재고 > 0:
            상태 = "🟠 주의" if 재고 <= reorder_threshold else "🟢 정상"
        else:
            상태 = determine_status(재고, 최근7일, 품절경과일, 조달유형 if 조달유형 != "미분류" else None)
        doi = compute_doi(재고, 최근7일)
        조치방안 = ACTION_BY_STATUS.get(상태, "")

        history_rows.append([
            target_str, 브랜드, code, 품목명, 상태, PRIORITY_BY_STATUS.get(상태, 99),
            전일재고, 재고, 출고, 입고계산, 최근7일, 최근90일, doi, 품절경과일, 조치방안,
        ])

        # 최근판매일/최근입고일 — 이력 전체(오늘 포함) 스캔. 날짜가 ISO(YYYY-MM-DD) 문자열이라
        # 문자열 비교가 곧 날짜 비교와 같다.
        최근판매일 = target_str if 출고 > 0 else None
        최근입고일 = target_str if _to_number(입고계산) > 0 else None
        for d_str, row in item_hist.items():
            if _to_number(row.get("출고")) > 0 and (최근판매일 is None or d_str > 최근판매일):
                최근판매일 = d_str
            if _to_number(row.get("입고")) > 0 and (최근입고일 is None or d_str > 최근입고일):
                최근입고일 = d_str
        미판매경과일 = (target_date - date.fromisoformat(최근판매일)).days if 최근판매일 else ""
        미입고경과일 = (target_date - date.fromisoformat(최근입고일)).days if 최근입고일 else ""

        # 입고단가/재고금액 — ITEM_COST_LOOKUP(2026-07-30 스냅샷)에 있는 품목만 계산되고,
        # 없으면 빈칸(모름을 0원으로 오판하면 안 됨). 악성재고/악성품절/샘플의심재고
        # 라우팅보다 앞에 있어야 셋 다 이 값을 쓸 수 있다.
        입고단가 = ITEM_COST_LOOKUP.get(code)
        재고금액 = round(재고 * 입고단가) if 입고단가 is not None else ""

        # ---- 3층 결과물 라우팅 ----
        # 디자인팀_발주필요는 "제작(자체제작) 발주"만 다루므로 조달유형이 자체제작인 품목만
        # 노출한다(2026-07-28 사용자 확정 — 국내사입/해외수입은 디자인팀이 아니라 관리팀 소관).
        is_design_target = 조달유형 == "자체제작"
        if 상태 == "⚫ 품절-장기":
            if 최근90일 > 0:
                if is_design_target:
                    design_rows.append([
                        브랜드, code, 품목명, 조달유형, 상태, PRIORITY_BY_STATUS.get(상태, 99),
                        리드타임, 재고, doi, 최근7일, 최근90일, 품절경과일,
                        "재입고 검토", f"재입고 골든타임(최근90일 {최근90일:.0f}개)",
                    ])
            elif 품절경과일 >= 90:
                # 90일 미만은 악성품절 대상에서 제외(2026-08-04 사용자 확정 — 30~89일짜리는
                # 관리팀_전체재고에 "품절-장기" 상태로는 계속 보이지만, 별도 검토 리스트인
                # 악성품절엔 진짜 오래 방치된 것만 올린다).
                maldead_rows.append([
                    브랜드, code, 품목명, 조달유형, 리드타임, 재고,
                    최근판매일 or "", 미판매경과일, 품절경과일, stockout_bucket(품절경과일),
                    최근입고일 or "", 미입고경과일, 최근90일,
                    "재발주 검토 또는 단종 검토 (판단 필요)", "",
                ])
        elif 상태 in PRIORITY_BY_STATUS and is_design_target:
            design_rows.append([
                브랜드, code, 품목명, 조달유형, 상태, PRIORITY_BY_STATUS.get(상태, 99),
                리드타임, 재고, doi, 최근7일, 최근90일, 품절경과일, 조치방안, "",
            ])

        # 관리팀_전체재고 — 전체 품목(필터 없음).
        wh_qtys = [by_item_wh.get(code, {}).get(wh, 0.0) for wh in OFFLINE_WAREHOUSES]
        mgmt_rows.append([
            브랜드, code, 품목명, 조달유형, 상태, 리드타임, *wh_qtys, 재고,
            전일재고, 입고계산, 출고, 최근7일, doi, 조치방안, "",
        ])

        # 악성재고 — 재고 있음 + 90일 이상 미판매.
        if 재고 > 0 and isinstance(미판매경과일, int) and 미판매경과일 >= 90:
            malstock_action = "프로모션/할인 검토" if 미판매경과일 < 180 else "땡처리/폐기 검토"
            malstock_rows.append([
                브랜드, code, 품목명, 재고, 최근판매일 or "", 미판매경과일,
                입고단가 if 입고단가 is not None else "", 재고금액, malstock_action, "",
            ])

        # 샘플의심재고 — 재고 1~2개 & 판매·입고 둘 다 3개월(90일) 이상 없음(2026-07-30 사용자
        # 확정 — 기존 60일에서 90일로 상향). 근거(미판매/미입고 둘 다 정수로 확정)가 없으면 아예
        # 후보에서 제외한다(도입 초기엔 이력이 짧아 근거가 없는 게 정상 — 근거 없음을 "이상
        # 없음"으로 오판하면 안 됨).
        if (재고 in (1, 2) and isinstance(미판매경과일, int) and 미판매경과일 >= 90
                and isinstance(미입고경과일, int) and 미입고경과일 >= 90):
            sample_rows.append([
                브랜드, code, 품목명, 조달유형, 재고, 최근판매일 or "", 미판매경과일,
                최근입고일 or "", 미입고경과일,
                입고단가 if 입고단가 is not None else "", 재고금액,
                "이카운트에서 샘플/전시용 여부 확인 필요", "",
            ])

    # 정렬 — 각 탭 note에 명시된 규칙.
    # POV_application(쇼핑백/봉투 등 비품, 판매용 아님)은 다른 판매 품목과 성격이 달라 섞이면
    # 헷갈리므로 맨 아래로 그룹핑(2026-07-28 사용자 확정) — 그 안에서는 기존 우선순위 정렬 유지.
    design_rows.sort(key=lambda r: (r[0] == "POV_application", r[5], r[7]))  # 비품 그룹 → 우선순위 asc → 재고수량 asc
    mgmt_rows.sort(key=lambda r: (PRIORITY_BY_STATUS.get(r[4], 50), r[1]))  # 상태우선순위 → 품목코드
    # 재고금액(r[7]) 내림차순 — 탭 note에 명시된 정렬 기준(2026-08-06부터 실제 값 채워짐).
    # 단가 근거 없어 재고금액이 빈칸("")인 행은 맨 뒤로.
    malstock_rows.sort(key=lambda r: -(r[7] if isinstance(r[7], (int, float)) else -1))
    maldead_rows.sort(key=lambda r: -(r[8] or 0))            # 품절경과일 내림차순
    sample_rows.sort(key=lambda r: -(r[6] or 0))              # 미판매(일) 내림차순

    return {
        "history": history_rows, "design": design_rows, "mgmt": mgmt_rows,
        "malstock": malstock_rows, "maldead": maldead_rows, "sample": sample_rows,
    }


# ---------------------------------------------------------------------------
# 5. 대시보드 — "오늘 처리할 것" 큐. 이미 계산된 4개 결과물 리스트를 잘라서 조립할 뿐,
#    새로 계산하지 않는다(같은 데이터가 두 군데서 다르게 나오는 걸 막기 위해).
# ---------------------------------------------------------------------------
DASHBOARD_HEADERS = ["순위", "브랜드", "품목명", "조달유형", "상태", "재고", "7일 판매", "DOI(소진일)", "경과(일)", "재고금액", "조치"]

# (제목, 배경색 톤) — 목업의 "색" 컬럼과 동일한 배정.
_BLOCK_TONE = {
    "danger": SEMANTIC_DANGER_BG, "info": SEMANTIC_INFO_BG, "warning": SEMANTIC_WARNING_BG, "dim": None,
}


def _dashboard_row(rank: int, brand, name, ptype, status, qty, sales7, doi, elapsed, money, action) -> list:
    return [rank, brand, name, ptype, status, qty, sales7, doi, elapsed, money, action]


def build_dashboard_blocks(result: dict) -> list[dict]:
    design, mgmt, malstock, maldead, sample = (
        result["design"], result["mgmt"], result["malstock"], result["maldead"], result["sample"],
    )

    def from_design(r, rank):
        # r: 브랜드,코드,품목명,조달유형,상태,우선순위,리드타임,재고,DOI,7일,90일,품절(일),조치,메모
        return _dashboard_row(rank, r[0], r[2], r[3], r[4], r[7], r[9], r[8], r[11], "", r[12])

    def from_mgmt(r, rank, elapsed=""):
        # r: 브랜드,코드,품목명,조달유형,상태,리드타임,4창고,총재고,전일재고,입고,출고,7일,DOI,조치,메모
        return _dashboard_row(rank, r[0], r[2], r[3], r[4], r[10], r[14], r[15], elapsed, "", r[16])

    def from_malstock(r, rank):
        # r: 브랜드,코드,품목명,재고,최근판매일,미판매(일),입고단가,재고금액,조치,메모
        # "과잉"(수량 많음)이 아니라 "미판매"(수량과 무관하게 오래 안 팔림)가 기준 —
        # 재고 1개짜리도 90일 이상 안 팔리면 여기 잡힌다(2026-08-05, 사용자 지적으로 라벨 수정).
        return _dashboard_row(rank, r[0], r[2], "", "🔵 미판매재고", r[3], "", "", r[5], r[7], r[8])

    def from_maldead(r, rank):
        # r: 브랜드,코드,품목명,조달유형,리드타임,재고,최근판매일,미판매(일),품절(일),품절구간,최근입고일,미입고(일),90일,조치,메모
        # 재고금액은 항상 0(품절 = 재고 0)이라 의미 없어서 안 넣는다.
        return _dashboard_row(rank, r[0], r[2], r[3], "⚫ 품절-장기", r[5], "", "", r[7], "", r[13])

    def from_sample(r, rank):
        # r: 브랜드,코드,품목명,조달유형,재고,최근판매일,미판매(일),최근입고일,미입고(일),입고단가,재고금액,조치,메모
        return _dashboard_row(rank, r[0], r[2], r[3], "🟣 샘플의심", r[4], "", "", r[6], r[10], r[11])

    blk1_rows = [from_design(r, i + 1) for i, r in enumerate(design[:20])]
    blk2_src = sorted([r for r in design if isinstance(r[11], int) and r[11] >= 3], key=lambda r: -r[11])
    blk2_rows = [from_design(r, i + 1) for i, r in enumerate(blk2_src[:10])]

    def out_of_stock_wh_count(r):
        return sum(1 for v in r[6:10] if v == 0)

    blk3_src = sorted(
        [r for r in mgmt if r[10] > 0 and out_of_stock_wh_count(r) > 0],
        key=lambda r: -out_of_stock_wh_count(r),
    )
    blk3_rows = [from_mgmt(r, i + 1) for i, r in enumerate(blk3_src[:20])]

    blk4_src = sorted(mgmt, key=lambda r: -(r[14] or 0))
    blk4_rows = [from_mgmt(r, i + 1) for i, r in enumerate(blk4_src[:10])]

    # 악성재고/악성품절은 "매일 처리할 목록"이 아니라 "주기적으로 훑어보는 감사 리스트"라 —
    # 상위 몇 개로 자르지 않고 전부 보여준다. 정렬 기준(재고금액/재고수량 등)도 강제하지 않고
    # daily_runner가 계산한 기본 정렬(미판매(일)/품절(일) 내림차순)만 유지, 나머지는 시트에서
    # 직접 정렬해서 보라는 사용자 요청 반영(2026-07-28).
    blk5_rows = [from_malstock(r, i + 1) for i, r in enumerate(malstock)]
    blk6_rows = [from_maldead(r, i + 1) for i, r in enumerate(maldead)]
    blk7_rows = [from_sample(r, i + 1) for i, r in enumerate(sample)]

    return [
        {"title": "오늘 조치 필요", "action": "지금 발주해야 리드타임 커버", "tone": "danger",
         "source": "디자인팀_발주필요", "sort": "우선순위 → 재고 오름차순",
         "total": len(design), "show": 20, "rows": blk1_rows},
        {"title": "3일 이상 품절", "action": "즉시 재입고 검토", "tone": "danger",
         "source": "디자인팀_발주필요", "sort": "품절(일) 내림차순",
         "total": len(blk2_src), "show": 10, "rows": blk2_rows},
        {"title": "창고이동 검토", "action": "결품 매장으로 재고 이동", "tone": "info",
         "source": "관리팀_전체재고 · 총재고>0", "sort": "결품 매장 수 내림차순",
         "total": len(blk3_src), "show": 20, "rows": blk3_rows},
        {"title": "판매 속도 상위", "action": "재입고 우선순위 확인", "tone": "warning",
         "source": "관리팀_전체재고", "sort": "최근 7일 판매량 내림차순",
         "total": len(mgmt), "show": 10, "rows": blk4_rows},
        {"title": "악성재고 정리", "action": "프로모션·번들·폐기 판단 (전체 목록, 필요시 시트에서 직접 정렬)", "tone": "info",
         "source": "악성재고", "sort": "미판매(일) 내림차순",
         "total": len(malstock), "show": len(malstock), "rows": blk5_rows},
        {"title": "악성품절 — 단종 판단", "action": "재발주 여부 · 단종 확정 (전체 목록, 필요시 시트에서 직접 정렬)", "tone": "dim",
         "source": "악성품절", "sort": "품절(일) 내림차순",
         "total": len(maldead), "show": len(maldead), "rows": blk6_rows},
        {"title": "샘플의심재고 확인", "action": "이카운트에서 샘플/전시용 여부 확인 (전체 목록, 필요시 시트에서 직접 정렬)", "tone": "info",
         "source": "샘플의심재고", "sort": "미판매(일) 내림차순",
         "total": len(sample), "show": len(sample), "rows": blk7_rows},
    ]


def build_status_distribution_line(mgmt_rows: list[list]) -> str:
    order = ["⛔️", "🔴", "🟠", "🟡", "🔵", "🟤", "⚫", "🟢"]
    counts = {c: 0 for c in order}
    for r in mgmt_rows:
        status = r[4] or ""
        for c in order:
            if status.startswith(c):
                counts[c] += 1
                break
    parts = " · ".join(f"{c} {counts[c]}" for c in order)
    return f"오늘 재고 현황 — 전체 {len(mgmt_rows)}개 — {parts}"


def write_status_distribution_banner(service, spreadsheet_id: str, mgmt_rows: list[list]) -> None:
    """관리팀_전체재고 2행(할 일 배너)을 상태 분포 한 줄로 매일 갱신한다."""
    line = build_status_distribution_line(mgmt_rows)
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id, range="'관리팀_전체재고'!A2",
        valueInputOption="RAW", body={"values": [[line]]},
    ).execute()


def with_money_total_row(rows: list[list], money_idx: int, ncols: int, label_idx: int = 2) -> list[list]:
    """헤더 바로 아래(맨 위 데이터 행)에 재고금액 합계 행을 하나 얹는다(2026-08-06 사용자 요청).

    새 리스트를 반환한다 — 원본 rows(예: result["malstock"])를 그대로 변형하면 대시보드가
    같은 리스트를 나중에 또 써서 합계 행이 개별 품목처럼 섞여 들어간다.
    """
    if not rows:
        return rows
    total = sum(r[money_idx] for r in rows if isinstance(r[money_idx], (int, float)))
    total_row = [""] * ncols
    total_row[label_idx] = f"▶ 합계 ({len(rows)}건)"
    total_row[money_idx] = total
    return [total_row] + rows


# 합계 행 배경색 — 사용자가 시트에서 직접 고른 색(2026-08-06, 채우기색 커스텀 HEX 확인).
TOTAL_ROW_BG_MALSTOCK = {"red": 1.0, "green": 0.949, "blue": 0.8}       # #fff2cc — 악성재고
TOTAL_ROW_BG_SAMPLE = {"red": 0.812, "green": 0.886, "blue": 0.953}    # #cfe2f3 — 샘플의심재고


def style_total_row(service, spreadsheet_id: str, sheet_id: int, ncols: int, bg: dict) -> None:
    """합계 행(맨 위 데이터 행)에 배경색 + 볼드를 매번 다시 칠한다.

    replace_tab_rows는 값만 지우고 다시 쓰기 때문에 서식 자체는 안 건드리지만,
    ecount_sheets_setup.py의 줄무늬(밴딩)를 재실행하면 밴딩 범위의 첫 줄(=이 행)이
    흰색으로 강제 지정되어 사용자가 수동으로 넣은 배경색이 사라진다(2026-08-06 확인).
    그래서 매일 실행마다 여기서 명시적으로 다시 칠해서 항상 유지되게 한다.
    """
    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": [{
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": DATA_START_IDX, "endRowIndex": DATA_START_IDX + 1,
                           "startColumnIndex": 0, "endColumnIndex": ncols},
                "cell": {"userEnteredFormat": {"backgroundColor": bg, "textFormat": {"bold": True, "fontFamily": FONT_BODY, "fontSize": 9}}},
                "fields": "userEnteredFormat(backgroundColor,textFormat)",
            }
        }]},
    ).execute()


def _status_style(status: str):
    """STATUS_COLOR_RULES를 재사용해 대시보드 상태 셀 색을 본문 탭과 동일하게 맞춘다."""
    for substr, bg, fg, bold in STATUS_COLOR_RULES:
        if substr in (status or ""):
            return bg, fg, bold
    return None, BODY_FG, False


def write_dashboard_tab(service, spreadsheet_id: str, sheet_id: int, blocks: list[dict], target_date: date) -> None:
    ncols = len(DASHBOARD_HEADERS)
    # 실행 시각과 기준일(데이터가 나타내는 날짜)을 둘 다 보여준다 — 이 시스템은 항상
    # "실행일-1일"을 처리하도록 설계돼있어(이카운트 판매현황 자동알림의 "전일" 관례를
    # 따름) 기준일만 보면 하루 늦은 것처럼 오해할 수 있다(2026-08-04 사용자 확인).
    run_at = datetime.now(KST)
    values: list[list] = [
        ["오늘 처리할 것을 위에서부터 순서대로", "", "", "", "", "", "", "", "", "", ""],
        [f"결과물 탭 5개에서 자동 집계 · {run_at.strftime('%Y-%m-%d %H:%M')}에 실행, "
         f"{target_date.isoformat()}(어제)까지의 ERP 데이터 반영", "", "", "", "", "", "", "", "", "", ""],
        [""] * ncols,
    ]
    row_styles: list[tuple[int, str, dict | None]] = []  # (row_idx0, kind, tone)
    MONEY_COL_IDX = DASHBOARD_HEADERS.index("재고금액")

    for block in blocks:
        count_label = f"전체 {block['total']}개" if block["show"] >= block["total"] else f"상위 {block['show']} / 전체 {block['total']}개"
        title_row = f"{block['title']} — {count_label} → {block['action']}"
        meta_row = f"출처: {block['source']} · 정렬: {block['sort']}"
        row_styles.append((len(values), "section", _BLOCK_TONE.get(block["tone"])))
        values.append([title_row] + [""] * (ncols - 2) + [meta_row])

        # 카테고리 제목 바로 아래에 재고금액 합계 — 값이 있는 블록(악성재고/샘플의심재고)만
        # 표시(2026-08-06 사용자 요청). 나머지 블록은 애초에 재고금액을 안 다뤄서 항상 0으로
        # 나와 오해를 줄 수 있어 생략.
        money_total = sum(r[MONEY_COL_IDX] for r in block["rows"] if isinstance(r[MONEY_COL_IDX], (int, float)))
        if money_total > 0:
            row_styles.append((len(values), "subtitle", None))
            values.append([f"재고금액 합계 {money_total:,.0f}원"] + [""] * (ncols - 1))

        row_styles.append((len(values), "header", None))
        values.append(list(DASHBOARD_HEADERS))
        status_rows = []
        for r in block["rows"]:
            status_rows.append(len(values))
            values.append(r)
        row_styles.append((-1, "statusrows", status_rows))  # 표시용, 실제 서식 루프에서 사용
        values.append([""] * ncols)  # 블록 사이 spacer

    service.spreadsheets().values().batchClear(
        spreadsheetId=spreadsheet_id, body={"ranges": ["'대시보드'!A1:Z2000"]}
    ).execute()
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id, range="'대시보드'!A1",
        valueInputOption="RAW", body={"values": values},
    ).execute()

    meta = service.spreadsheets().get(
        spreadsheetId=spreadsheet_id, fields="sheets(properties(sheetId,title),bandedRanges(bandedRangeId))"
    ).execute()
    delete_requests = []
    for s in meta["sheets"]:
        if s["properties"]["sheetId"] != sheet_id:
            continue
        delete_requests.append({"unmergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": len(values), "startColumnIndex": 0, "endColumnIndex": ncols}}})
        for banded in s.get("bandedRanges", []):
            delete_requests.append({"deleteBanding": {"bandedRangeId": banded["bandedRangeId"]}})
    if delete_requests:
        service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": delete_requests}).execute()

    requests = [{
        "repeatCell": {
            "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": len(values), "startColumnIndex": 0, "endColumnIndex": ncols},
            "cell": {"userEnteredFormat": {"textFormat": {"fontFamily": FONT_BODY, "fontSize": 9, "foregroundColor": BODY_FG}}},
            "fields": "userEnteredFormat.textFormat",
        }
    }]
    # 배너 2행.
    requests.append({"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": ncols}, "mergeType": "MERGE_ALL"}})
    requests.append({"repeatCell": {"range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": ncols},
                                     "cell": {"userEnteredFormat": {"backgroundColor": BANNER1_BG, "textFormat": {"bold": True, "foregroundColor": BANNER_FG_LIGHT, "fontFamily": FONT_BODY, "fontSize": 11}, "horizontalAlignment": "LEFT", "verticalAlignment": "MIDDLE"}},
                                     "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)"}})
    requests.append({"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 2, "startColumnIndex": 0, "endColumnIndex": ncols}, "mergeType": "MERGE_ALL"}})
    requests.append({"repeatCell": {"range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 2, "startColumnIndex": 0, "endColumnIndex": ncols},
                                     "cell": {"userEnteredFormat": {"backgroundColor": BANNER3_BG, "textFormat": {"foregroundColor": BANNER3_FG, "fontFamily": FONT_BODY, "fontSize": 9}}},
                                     "fields": "userEnteredFormat(backgroundColor,textFormat)"}})

    for row_idx, kind, extra in row_styles:
        if kind == "section":
            tone_bg = extra
            requests.append({"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": row_idx, "endRowIndex": row_idx + 1, "startColumnIndex": 0, "endColumnIndex": ncols - 6}, "mergeType": "MERGE_ALL"}})
            requests.append({"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": row_idx, "endRowIndex": row_idx + 1, "startColumnIndex": ncols - 6, "endColumnIndex": ncols}, "mergeType": "MERGE_ALL"}})
            fmt = {"textFormat": {"bold": True, "fontFamily": FONT_BODY, "fontSize": 10}}
            if tone_bg is not None:
                fmt["backgroundColor"] = tone_bg
            requests.append({"repeatCell": {"range": {"sheetId": sheet_id, "startRowIndex": row_idx, "endRowIndex": row_idx + 1, "startColumnIndex": 0, "endColumnIndex": ncols}, "cell": {"userEnteredFormat": fmt}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}})
        elif kind == "subtitle":
            requests.append({"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": row_idx, "endRowIndex": row_idx + 1, "startColumnIndex": 0, "endColumnIndex": ncols}, "mergeType": "MERGE_ALL"}})
            requests.append({"repeatCell": {"range": {"sheetId": sheet_id, "startRowIndex": row_idx, "endRowIndex": row_idx + 1, "startColumnIndex": 0, "endColumnIndex": ncols},
                                             "cell": {"userEnteredFormat": {"textFormat": {"italic": True, "foregroundColor": BODY_FG, "fontFamily": FONT_BODY, "fontSize": 9}, "horizontalAlignment": "RIGHT"}},
                                             "fields": "userEnteredFormat(textFormat,horizontalAlignment)"}})
        elif kind == "header":
            requests.append({"repeatCell": {"range": {"sheetId": sheet_id, "startRowIndex": row_idx, "endRowIndex": row_idx + 1, "startColumnIndex": 0, "endColumnIndex": ncols},
                                             "cell": {"userEnteredFormat": {"backgroundColor": HEADER_BG, "textFormat": {"bold": True, "foregroundColor": HEADER_FG, "fontFamily": FONT_BODY, "fontSize": 9}}},
                                             "fields": "userEnteredFormat(backgroundColor,textFormat)"}})
        elif kind == "statusrows":
            for r_idx in extra:
                status_val = values[r_idx][4]
                bg, fg, bold = _status_style(status_val)
                fmt = {"textFormat": {"foregroundColor": fg, "bold": bold, "fontFamily": FONT_BODY, "fontSize": 9}}
                if bg is not None:
                    fmt["backgroundColor"] = bg
                requests.append({"repeatCell": {"range": {"sheetId": sheet_id, "startRowIndex": r_idx, "endRowIndex": r_idx + 1, "startColumnIndex": 4, "endColumnIndex": 5}, "cell": {"userEnteredFormat": fmt}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}})
                # 숫자 컬럼(재고/7일판매/DOI/경과/재고금액) 고정폭 우측정렬.
                requests.append({"repeatCell": {"range": {"sheetId": sheet_id, "startRowIndex": r_idx, "endRowIndex": r_idx + 1, "startColumnIndex": 5, "endColumnIndex": 10},
                                                 "cell": {"userEnteredFormat": {"horizontalAlignment": "RIGHT", "textFormat": {"fontFamily": FONT_MONO, "fontSize": 9}}},
                                                 "fields": "userEnteredFormat(horizontalAlignment,textFormat)"}})
                # 재고는 관리팀_전체재고의 총재고와 동일하게 항상 볼드체로 강조(2026-07-28 사용자 요청).
                requests.append({"repeatCell": {"range": {"sheetId": sheet_id, "startRowIndex": r_idx, "endRowIndex": r_idx + 1, "startColumnIndex": 5, "endColumnIndex": 6},
                                                 "cell": {"userEnteredFormat": {"textFormat": {"bold": True, "fontFamily": FONT_MONO, "fontSize": 9}}},
                                                 "fields": "userEnteredFormat.textFormat"}})
                # 재고금액 — 콤마 서식 안 넣으면 자릿수가 안 읽힌다(2026-08-06 사용자 지적).
                requests.append({"repeatCell": {"range": {"sheetId": sheet_id, "startRowIndex": r_idx, "endRowIndex": r_idx + 1, "startColumnIndex": MONEY_COL_IDX, "endColumnIndex": MONEY_COL_IDX + 1},
                                                 "cell": {"userEnteredFormat": {"numberFormat": NUMBER_FORMAT_MONEY}},
                                                 "fields": "userEnteredFormat.numberFormat"}})

    requests.append({"updateSheetProperties": {"properties": {"sheetId": sheet_id, "gridProperties": {"frozenRowCount": 2}}, "fields": "gridProperties.frozenRowCount"}})

    try:
        service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": requests}).execute()
    except Exception as e:
        print(f"[runner] 대시보드 서식 적용 중 일부 실패(무시하고 진행): {e}")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--spreadsheet-id", required=True)
    ap.add_argument("--base-date", help="YYYY-MM-DD (기본: 실행일 기준 어제, KST)")
    ap.add_argument("--sales-xlsx", help="판매현황 엑셀 경로를 수동 지정(테스트용). 미지정 시 기본 동작인 "
                                          "웹 스크래핑으로 대체됨 — 자동알림 메일에는 더 이상 첨부가 없음(2026-07-28 확인)")
    ap.add_argument("--gmail-query", default='from:ecountnotice@ecount.com has:attachment newer_than:2d',
                     help="--sales-xlsx 없이 엑셀 첨부로 받던 옛 경로에서만 쓰임(현재는 미사용, 하위호환용)")
    ap.add_argument("--dry-run", action="store_true", help="계산만 하고 시트에 쓰지 않음")
    ap.add_argument("--use-cached-inventory", action="store_true",
                     help="직전 실행에서 저장된 재고API 캐시를 재사용(창고당 10분 대기 없이 즉시) — "
                          "재고 조회는 성공했는데 그 다음 단계(웹 스크래핑 등)에서 실패해 재시도할 때 유용")
    ap.add_argument("--use-cached-sales", action="store_true",
                     help="직전 실행에서 저장된 판매현황 스크래핑 캐시를 재사용(Gmail/로그인 없이 즉시) — "
                          "스크래핑은 성공했는데 그 다음 단계에서 실패해 재시도할 때 유용")
    args = ap.parse_args()

    target_date = date.fromisoformat(args.base_date) if args.base_date else (datetime.now(KST).date() - timedelta(days=1))
    base_date_str = target_date.strftime("%Y%m%d")
    print(f"[runner] 기준일(TARGET_DATE) = {target_date.isoformat()}")

    inventory_cache_path = DUMP_DIR / f"inventory_raw_{target_date.isoformat()}.json"
    if args.use_cached_inventory:
        if not inventory_cache_path.exists():
            raise SystemExit(f"[runner] 캐시 파일이 없습니다: {inventory_cache_path} — --use-cached-inventory 없이 다시 실행하세요.")
        inventory_raw = json.loads(inventory_cache_path.read_text())
        print(f"[runner] 재고 원본 캐시 사용: {inventory_cache_path} ({len(inventory_raw)}건)")
    else:
        print("[runner] 이카운트 재고API 조회 중...")
        inventory_raw = fetch_inventory_raw(base_date_str)
        print(f"[runner] 재고 원본 {len(inventory_raw)}건 (전체 창고)")
        inventory_cache_path.parent.mkdir(parents=True, exist_ok=True)
        inventory_cache_path.write_text(json.dumps(inventory_raw, ensure_ascii=False))
        print(f"[runner] 재고 원본 캐시 저장: {inventory_cache_path} (다음 실행에서 --use-cached-inventory로 재사용 가능)")

    sales_cache_path = DUMP_DIR / f"sales_raw_{target_date.isoformat()}.json"
    if args.sales_xlsx:
        sales_raw = parse_sales_xlsx(Path(args.sales_xlsx))
        print(f"[runner] 판매 원본 {len(sales_raw)}건 (전체 창고, 출처={args.sales_xlsx})")
    elif args.use_cached_sales:
        if not sales_cache_path.exists():
            raise SystemExit(f"[runner] 캐시 파일이 없습니다: {sales_cache_path} — --use-cached-sales 없이 다시 실행하세요.")
        sales_raw = json.loads(sales_cache_path.read_text())
        print(f"[runner] 판매 원본 캐시 사용: {sales_cache_path} ({len(sales_raw)}건)")
    else:
        print("[runner] 이카운트 판매현황 웹 리포트 스크래핑 중... (Gmail 링크 → 로그인 → 표 읽기)")
        from ecount_sales_scraper import scrape_sales_status
        sales_raw = scrape_sales_status()
        print(f"[runner] 판매 원본 {len(sales_raw)}건 (전체 창고, 웹 스크래핑)")
        sales_cache_path.parent.mkdir(parents=True, exist_ok=True)
        sales_cache_path.write_text(json.dumps(sales_raw, ensure_ascii=False))
        print(f"[runner] 판매 원본 캐시 저장: {sales_cache_path} (다음 실행에서 --use-cached-sales로 재사용 가능)")

    creds = _load_creds()
    from googleapiclient.discovery import build
    service = build("sheets", "v4", credentials=creds)

    print("[runner] 품목마스터/이력 로드 중...")
    master_headers = TABS["품목마스터"]["headers"]
    master_rows = read_tab_rows(service, args.spreadsheet_id, "품목마스터")
    item_master = {
        r["품목코드"]: {"브랜드": r["브랜드"], "조달유형": r["조달유형"],
                       "리드타임": _to_number(r["리드타임(일)"]) if r["리드타임(일)"] else "",
                       "품목명": r["품목명"]}
        for r in master_rows if r["품목코드"]
    }
    history = read_tab_rows(service, args.spreadsheet_id, "일별재고이력")
    print(f"[runner] 품목마스터 {len(item_master)}건, 이력 {len(history)}행 로드")

    new_master_rows = auto_register_new_items(item_master, sales_raw, target_date)
    if new_master_rows:
        unclassified = sum(1 for r in new_master_rows if r[4] == "미분류")
        print(f"[runner] 품목마스터 신규 자동등록 {len(new_master_rows)}건"
              + (f" (이 중 {unclassified}건은 새 브랜드라 미분류 — 품목마스터에서 조달유형 확인 필요)" if unclassified else ""))

    result = build_daily_rows(target_date, inventory_raw, sales_raw, item_master, history)
    print(f"[runner] 계산 완료 — 일별이력 {len(result['history'])}건 / 디자인팀_발주필요 "
          f"{len(result['design'])}건 / 관리팀_전체재고 {len(result['mgmt'])}건 / "
          f"악성재고 {len(result['malstock'])}건 / 악성품절 {len(result['maldead'])}건 / "
          f"샘플의심재고 {len(result['sample'])}건")

    if args.dry_run:
        print("[runner] --dry-run — 시트에 쓰지 않음")
        return 0

    print("[runner] RAW 탭 반영 중...")
    raw_inv_rows = [[r[h] for h in TABS["RAW_재고현황"]["headers"]] for r in inventory_raw]
    raw_sales_rows = [[r[h] for h in TABS["RAW_판매현황"]["headers"]] for r in sales_raw]
    replace_tab_rows(service, args.spreadsheet_id, "RAW_재고현황", raw_inv_rows)
    replace_tab_rows(service, args.spreadsheet_id, "RAW_판매현황", raw_sales_rows)

    print("[runner] 일별재고이력 반영 중...")
    append_history_rows(service, args.spreadsheet_id, target_date.isoformat(), result["history"])

    print("[runner] 3층 결과물 탭 반영 중...")
    replace_tab_rows(service, args.spreadsheet_id, "디자인팀_발주필요", result["design"])
    replace_tab_rows(service, args.spreadsheet_id, "관리팀_전체재고", result["mgmt"])
    replace_tab_rows(service, args.spreadsheet_id, "악성재고",
                      with_money_total_row(result["malstock"], money_idx=7, ncols=10))
    replace_tab_rows(service, args.spreadsheet_id, "악성품절", result["maldead"])
    replace_tab_rows(service, args.spreadsheet_id, "샘플의심재고",
                      with_money_total_row(result["sample"], money_idx=10, ncols=13))
    write_status_distribution_banner(service, args.spreadsheet_id, result["mgmt"])

    tab_meta = service.spreadsheets().get(
        spreadsheetId=args.spreadsheet_id, fields="sheets.properties(sheetId,title)"
    ).execute()
    tab_id_by_title = {s["properties"]["title"]: s["properties"]["sheetId"] for s in tab_meta["sheets"]}
    if "악성재고" in tab_id_by_title:
        style_total_row(service, args.spreadsheet_id, tab_id_by_title["악성재고"], ncols=10, bg=TOTAL_ROW_BG_MALSTOCK)
    if "샘플의심재고" in tab_id_by_title:
        style_total_row(service, args.spreadsheet_id, tab_id_by_title["샘플의심재고"], ncols=13, bg=TOTAL_ROW_BG_SAMPLE)

    if new_master_rows:
        print("[runner] 품목마스터 신규 품목 반영 중...")
        append_item_master_rows(service, args.spreadsheet_id, new_master_rows)

    print("[runner] 대시보드 반영 중...")
    dash_meta = service.spreadsheets().get(
        spreadsheetId=args.spreadsheet_id, fields="sheets.properties(sheetId,title)"
    ).execute()
    dash_id_by_title = {s["properties"]["title"]: s["properties"]["sheetId"] for s in dash_meta["sheets"]}
    if "대시보드" in dash_id_by_title:
        blocks = build_dashboard_blocks(result)
        write_dashboard_tab(service, args.spreadsheet_id, dash_id_by_title["대시보드"], blocks, target_date)
    else:
        print("[runner] ⚠️ '대시보드' 탭이 없습니다 — ecount_sheets_setup.py를 먼저 실행하세요.")

    print("[runner] 완료.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
